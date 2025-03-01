import re
import os
import pandas as pd
import openpyxl
import requests
from rest_framework import status
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework.views import APIView
from django.conf import settings
from common.clean import clean_product_text
from PyPDF2 import PdfReader, PdfWriter, PdfMerger
from io import BytesIO


class HandlerFile(APIView):
    def post(self, request: Request) -> Response:
        files_upload = request.FILES
        merges = PdfMerger()
        HOST_PARSER = settings.HOST_PARSER
        try:
            pages : list[tuple] = []
            for index_files, files in enumerate(files_upload):
                reader = PdfReader(files_upload[files])
                for index_page, page in enumerate(reader.pages):    
                    writer = PdfWriter()
                    writer.add_page(page)
                    pdf_buffer = BytesIO()
                    writer.write(pdf_buffer)
                    pages.append(('files[]', (F"{index_files}_{index_page}.pdf", pdf_buffer.getvalue(), 'application/pdf')))
                    merges.append(pdf_buffer)
            
            url = F"http://ocr-api:{HOST_PARSER}/pdf-parser"
            response = requests.request("GET", url, files=pages)
            responseText : list[dict] = response.json()
            orders: list = []
            
            for index_page, page in enumerate(responseText):
                contentOrder = page['ContentOrder']
                _product_lines = re.split(r'(?=\d+\.)', contentOrder)
                product_lines = [pro.strip() for pro in _product_lines if pro != ""]
                for index_product, product in enumerate(product_lines):
                    content_order, sl = clean_product_text(product)
                    orders.append({
                        "page" : index_page,
                        "order_id" : page["OrderId"],
                        "content": content_order,
                        "quantity": sl,
                        "total_product": len(product_lines),
                    })
                    
                    
            # created orders
            # get bytes data from merges pdf
            allpdf_buffer = BytesIO()
            merges.write(allpdf_buffer)
            
            return Response(data = {
                "orders" : orders,
                "pdf" : str(allpdf_buffer.getvalue())
                }, status = status.HTTP_200_OK)
            
            
        except Exception as error:
            print(error)
            return Response(data = {}, status = status.HTTP_400_BAD_REQUEST)
        
        
class HandlerSplitSKU(APIView):
    
    def post(self, request: Request) -> Response:
        try:
            GUIDE_PATH = settings.GUIDE_PATH
            orders = request.data['orders']
            orders_sheets = openpyxl.Workbook()
            sheet_description = orders_sheets.active
            sheet_description.title = "Orders"
            
            sheet_guide = orders_sheets.create_sheet("GUIDE")
            sheet_guide.title = "GUIDE"            
            
            sheet_description.append([
                "Page",
                "Order ID",
                "Content",
                "Quantity",
                "Total Product",
            ])
            for index_order, order in enumerate(orders):
                sheet_description.append([
                    index_order + 1,
                    order['order_id'],
                    order['content'].replace("\n", "").replace("\r", "").replace(" ", ""),
                    order['quantity'],
                    order['total_product'],
                ])
            wb_guide = openpyxl.load_workbook(GUIDE_PATH)
            rows_guide = []
            for row in wb_guide["GUIDE"].iter_rows():
                sheet_guide.append([cell.value for cell in row])
                rows_guide.append([cell.value for cell in row])
                
            # save orders_sheets
            # orders_sheets.save("orders.xlsx")
            guide_data : dict = dict()
            for index_row, row in enumerate(rows_guide):
                if index_row == 0:
                    continue
                key = str(row[0]).strip() if row[0] else ""
                value = row[2] if len(row) > 2 and row[2] else ""
                guide_data[key] = value
                
                
            sheet_description.cell(row = 1, column = 6, value = "TSP")
                
            # # Convert sheet_description data to DataFrame
            data = sheet_description.values
            
            orders_sheets.save("orders.xlsx")
            
            
            
            # for order in orders:
                # print(order)
                
            return Response(data={}, status = status.HTTP_200_OK)
        except Exception as error:
            print(error)
            return Response(data = {}, status = status.HTTP_400_BAD_REQUEST)